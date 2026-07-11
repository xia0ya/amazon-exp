import json
import math
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent
ABA_FILE_PATTERN = "ABA关键词趋势分析报表*.xlsx"
OUTPUT_DIR = ROOT / "outputs"
OUTPUT_FILE = OUTPUT_DIR / "keyword_trend_comparison.html"


def clean_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if text in {"", "-", "未进前三"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m", "%Y/%m"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.date()
        except ValueError:
            pass
    return None


def read_table(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    records = []
    for row in rows[1:]:
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue
        rec = {}
        for idx, header in enumerate(headers):
            if header:
                rec[header] = row[idx] if idx < len(row) else None
        records.append(rec)
    return records


def short_title(title):
    if not title:
        return ""
    text = re.sub(r"&amp;", "&", str(title))
    text = re.sub(r"\s+", " ", text).strip()
    return text[:92] + ("..." if len(text) > 92 else "")


def asin_from_filename(path):
    match = re.search(r"(B0[A-Z0-9]{8}|B[A-Z0-9]{9})", path.name)
    return match.group(1) if match else path.stem


def load_aba():
    matches = sorted(ROOT.glob(ABA_FILE_PATTERN))
    if not matches:
        raise FileNotFoundError("No ABA keyword trend workbook found.")
    path = matches[0]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    trend_rows = []
    if "排名趋势" in wb.sheetnames:
        for rec in read_table(wb["排名趋势"]):
            start = parse_date(rec.get("数据开始日期"))
            end = parse_date(rec.get("数据结束日期"))
            rank = clean_number(rec.get("搜索量排名"))
            if start and end and rank is not None:
                trend_rows.append(
                    {
                        "country": str(rec.get("国家") or ""),
                        "keyword": str(rec.get("关键词") or ""),
                        "start": start.isoformat(),
                        "end": end.isoformat(),
                        "rank": int(rank),
                        "month": end.strftime("%Y-%m"),
                    }
                )

    top_asins = []
    if "前三ASIN" in wb.sheetnames:
        for rec in read_table(wb["前三ASIN"]):
            asin = str(rec.get("asin") or "").strip()
            if not asin:
                continue
            latest_raw = rec.get("最新商品排位")
            latest_rank = None if str(latest_raw).strip() == "未进前三" else clean_number(latest_raw)
            top_asins.append(
                {
                    "asin": asin,
                    "title": short_title(rec.get("标题")),
                    "latestRank": int(latest_rank) if latest_rank is not None else None,
                    "totalPeriods": int(clean_number(rec.get("总周期数")) or 0),
                    "rank1Periods": int(clean_number(rec.get("榜一期数")) or 0),
                    "top3Periods": int(clean_number(rec.get("前三期数")) or 0),
                }
            )
    return path.name, sorted(trend_rows, key=lambda x: x["end"]), top_asins


def load_sales():
    sales_files = [
        p
        for p in sorted(ROOT.glob("*.xlsx"))
        if not p.name.startswith("ABA关键词趋势分析报表")
    ]
    products = {}
    monthly_rows = []
    daily_rows = []

    for path in sales_files:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        asin = asin_from_filename(path)
        product = {
            "asin": asin,
            "file": path.name,
            "label": path.stem,
            "title": "",
            "latestMonth": None,
            "latestMonthSales": 0,
            "latestMonthRevenue": 0,
        }

        for ws in wb.worksheets:
            if "月销量" in ws.title:
                for rec in read_table(ws):
                    month = str(rec.get("最近几月") or "").strip()
                    row_asin = str(rec.get("ASIN") or asin).strip()
                    sales = clean_number(rec.get("月销量"))
                    revenue = clean_number(rec.get("月销售额($)"))
                    price = clean_number(rec.get("平均单价($)"))
                    title = short_title(rec.get("商品"))
                    if title and not product["title"]:
                        product["title"] = title
                    if month and sales is not None:
                        item = {
                            "asin": row_asin,
                            "month": month,
                            "sales": sales,
                            "revenue": revenue or 0,
                            "price": price or 0,
                        }
                        monthly_rows.append(item)
                        if product["latestMonth"] is None or month > product["latestMonth"]:
                            product["latestMonth"] = month
                            product["latestMonthSales"] = sales
                            product["latestMonthRevenue"] = revenue or 0
            elif "日销量" in ws.title:
                for rec in read_table(ws):
                    day = parse_date(rec.get("时间"))
                    row_asin = str(rec.get("ASIN") or asin).strip()
                    sales = clean_number(rec.get("日销量"))
                    if day and sales is not None:
                        daily_rows.append({"asin": row_asin, "date": day.isoformat(), "sales": sales})

        products[asin] = product

    return list(products.values()), monthly_rows, daily_rows


def aggregate_weekly_sales(trend_rows, daily_rows):
    by_asin_day = defaultdict(dict)
    for row in daily_rows:
        by_asin_day[row["asin"]][row["date"]] = row["sales"]

    weekly = []
    for trend in trend_rows:
        start = parse_date(trend["start"])
        end = parse_date(trend["end"])
        by_asin = {}
        total = 0.0
        for asin, days in by_asin_day.items():
            value = 0.0
            for day_text, sales in days.items():
                day = parse_date(day_text)
                if start <= day <= end:
                    value += sales
            if value > 0:
                by_asin[asin] = round(value, 2)
                total += value
        weekly.append(
            {
                "start": trend["start"],
                "end": trend["end"],
                "month": trend["month"],
                "rank": trend["rank"],
                "totalSales": round(total, 2),
                "asinSales": by_asin,
            }
        )
    return weekly


def aggregate_monthly(monthly_rows, trend_rows):
    sales_by_month = defaultdict(lambda: {"sales": 0.0, "revenue": 0.0, "asinCount": 0})
    seen = defaultdict(set)
    for row in monthly_rows:
        month = row["month"]
        sales_by_month[month]["sales"] += row["sales"]
        sales_by_month[month]["revenue"] += row["revenue"]
        seen[month].add(row["asin"])
    for month, asins in seen.items():
        sales_by_month[month]["asinCount"] = len(asins)

    ranks = defaultdict(list)
    for row in trend_rows:
        ranks[row["month"]].append(row["rank"])

    months = sorted(set(sales_by_month) | set(ranks))
    result = []
    for month in months:
        rank_values = ranks.get(month, [])
        sales = sales_by_month.get(month, {"sales": 0, "revenue": 0, "asinCount": 0})
        result.append(
            {
                "month": month,
                "avgRank": round(sum(rank_values) / len(rank_values), 1) if rank_values else None,
                "sales": round(sales["sales"], 2),
                "revenue": round(sales["revenue"], 2),
                "asinCount": sales["asinCount"],
            }
        )
    return result


def correlation(points, x_key, y_key):
    pairs = [(p[x_key], p[y_key]) for p in points if p.get(x_key) is not None and p.get(y_key) not in (None, 0)]
    if len(pairs) < 3:
        return None
    xs, ys = zip(*pairs)
    mx = sum(xs) / len(xs)
    my = sum(ys) / len(ys)
    num = sum((x - mx) * (y - my) for x, y in pairs)
    den_x = math.sqrt(sum((x - mx) ** 2 for x in xs))
    den_y = math.sqrt(sum((y - my) ** 2 for y in ys))
    if den_x == 0 or den_y == 0:
        return None
    return round(num / (den_x * den_y), 3)


def html_template(data):
    payload = json.dumps(data, ensure_ascii=False)
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>关键词趋势与销量对比分析</title>
  <style>
    :root {{
      --bg: #f6f7f9;
      --panel: #ffffff;
      --ink: #17212b;
      --muted: #627082;
      --line: #dde3ea;
      --accent: #1464d2;
      --accent2: #d65f32;
      --accent3: #16845b;
      --warn: #a86a00;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: Arial, "Microsoft YaHei", sans-serif;
      color: var(--ink);
      background: var(--bg);
    }}
    header {{
      padding: 26px 30px 18px;
      background: #fff;
      border-bottom: 1px solid var(--line);
    }}
    h1 {{ margin: 0 0 8px; font-size: 26px; letter-spacing: 0; }}
    h2 {{ margin: 0 0 14px; font-size: 18px; letter-spacing: 0; }}
    .meta {{ color: var(--muted); font-size: 13px; line-height: 1.6; }}
    main {{ padding: 22px 30px 32px; max-width: 1500px; margin: 0 auto; }}
    .kpis {{
      display: grid;
      grid-template-columns: repeat(5, minmax(150px, 1fr));
      gap: 12px;
      margin-bottom: 18px;
    }}
    .kpi, section {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
    }}
    .kpi {{ padding: 14px 16px; min-height: 86px; }}
    .kpi .label {{ color: var(--muted); font-size: 12px; margin-bottom: 8px; }}
    .kpi .value {{ font-size: 24px; font-weight: 700; }}
    .kpi .sub {{ color: var(--muted); font-size: 12px; margin-top: 6px; line-height: 1.35; }}
    .grid {{
      display: grid;
      grid-template-columns: minmax(0, 1.45fr) minmax(380px, .9fr);
      gap: 16px;
      align-items: start;
    }}
    section {{ padding: 16px; margin-bottom: 16px; }}
    .wide {{ grid-column: 1 / -1; }}
    .controls {{ display: flex; gap: 10px; flex-wrap: wrap; align-items: center; margin-bottom: 12px; }}
    .controls label {{ font-size: 13px; color: var(--muted); }}
    select, input {{
      border: 1px solid var(--line);
      background: #fff;
      border-radius: 6px;
      padding: 7px 9px;
      font-size: 13px;
      color: var(--ink);
    }}
    .chart {{ width: 100%; height: 360px; border: 1px solid var(--line); border-radius: 6px; background: #fff; }}
    .chart.small {{ height: 300px; }}
    svg {{ display: block; width: 100%; height: 100%; }}
    .legend {{ display: flex; gap: 16px; flex-wrap: wrap; color: var(--muted); font-size: 12px; margin-top: 10px; }}
    .dot {{ display: inline-block; width: 10px; height: 10px; border-radius: 50%; margin-right: 6px; vertical-align: -1px; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    th, td {{ border-bottom: 1px solid var(--line); padding: 8px 7px; text-align: left; vertical-align: top; }}
    th {{ color: var(--muted); font-weight: 700; background: #fafbfc; position: sticky; top: 0; }}
    .table-wrap {{ max-height: 420px; overflow: auto; border: 1px solid var(--line); border-radius: 6px; }}
    .tag {{ display: inline-block; padding: 2px 6px; border-radius: 999px; background: #eef4ff; color: #164f9f; font-size: 12px; }}
    .note {{ color: var(--muted); font-size: 12px; line-height: 1.5; margin-top: 10px; }}
    .tooltip {{
      position: fixed;
      pointer-events: none;
      opacity: 0;
      background: rgba(23,33,43,.94);
      color: #fff;
      padding: 8px 10px;
      border-radius: 6px;
      font-size: 12px;
      line-height: 1.45;
      max-width: 320px;
      z-index: 9;
    }}
    @media (max-width: 980px) {{
      header, main {{ padding-left: 16px; padding-right: 16px; }}
      .kpis {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
      .grid {{ grid-template-columns: 1fr; }}
      .chart {{ height: 320px; }}
    }}
  </style>
</head>
<body>
  <header>
    <h1>关键词趋势与销量对比分析</h1>
    <div class="meta" id="meta"></div>
  </header>
  <main>
    <div class="kpis" id="kpis"></div>
    <div class="grid">
      <section>
        <h2>ABA 搜索排名 vs 汇总周销量</h2>
        <div class="chart" id="weeklyChart"></div>
        <div class="legend">
          <span><i class="dot" style="background:var(--accent)"></i>搜索量排名，越靠上越好</span>
          <span><i class="dot" style="background:var(--accent2)"></i>所有已导入 ASIN 周销量</span>
        </div>
        <div class="note">周销量按 ABA 的数据开始/结束日期，对每个 Sales 文件的日销量求和。</div>
      </section>
      <section>
        <h2>关键词前三 ASIN 占位</h2>
        <div class="chart small" id="topAsinChart"></div>
        <div class="note">“榜一期数 / 前三期数”来自 ABA 报表的前三 ASIN sheet。</div>
      </section>

      <section class="wide">
        <h2>月度对比</h2>
        <div class="controls">
          <label>销量指标
            <select id="monthlyMetric">
              <option value="sales">月销量</option>
              <option value="revenue">月销售额($)</option>
            </select>
          </label>
        </div>
        <div class="chart" id="monthlyChart"></div>
        <div class="legend">
          <span><i class="dot" style="background:var(--accent)"></i>ABA 月均搜索排名，越靠上越好</span>
          <span><i class="dot" style="background:var(--accent3)"></i>Sales 表月度汇总</span>
        </div>
      </section>

      <section>
        <h2>ASIN 月销量排行</h2>
        <div class="controls">
          <label>月份
            <select id="rankMonth"></select>
          </label>
        </div>
        <div class="chart small" id="asinBarChart"></div>
      </section>
      <section>
        <h2>单个 ASIN 周销量联动</h2>
        <div class="controls">
          <label>ASIN
            <select id="asinSelect"></select>
          </label>
        </div>
        <div class="chart small" id="asinWeeklyChart"></div>
      </section>

      <section class="wide">
        <h2>明细表</h2>
        <div class="controls">
          <label>搜索 ASIN/文件
            <input id="productFilter" placeholder="输入 ASIN 或文件名" />
          </label>
        </div>
        <div class="table-wrap">
          <table id="productTable"></table>
        </div>
      </section>
    </div>
  </main>
  <div class="tooltip" id="tooltip"></div>
  <script>
    const DATA = {payload};
    const fmt = new Intl.NumberFormat('en-US');
    const money = new Intl.NumberFormat('en-US', {{ style: 'currency', currency: 'USD', maximumFractionDigits: 0 }});
    const tooltip = document.getElementById('tooltip');

    function showTip(html, event) {{
      tooltip.innerHTML = html;
      tooltip.style.opacity = 1;
      tooltip.style.left = Math.min(event.clientX + 14, window.innerWidth - 340) + 'px';
      tooltip.style.top = (event.clientY + 14) + 'px';
    }}
    function hideTip() {{ tooltip.style.opacity = 0; }}
    function val(v) {{ return v == null ? '—' : fmt.format(Math.round(v)); }}

    function metricRange(values) {{
      const nums = values.filter(v => Number.isFinite(v));
      if (!nums.length) return [0, 1];
      let min = Math.min(...nums), max = Math.max(...nums);
      if (min === max) {{ min -= 1; max += 1; }}
      return [min, max];
    }}
    function makeSvg(containerId) {{
      const el = document.getElementById(containerId);
      el.innerHTML = '';
      const svg = document.createElementNS('http://www.w3.org/2000/svg', 'svg');
      svg.setAttribute('viewBox', '0 0 1000 420');
      el.appendChild(svg);
      return svg;
    }}
    function linePath(points) {{
      return points.map((p, i) => `${{i ? 'L' : 'M'}}${{p.x.toFixed(1)}},${{p.y.toFixed(1)}}`).join(' ');
    }}
    function drawAxes(svg, labels, leftTitle, rightTitle) {{
      const grid = document.createElementNS('http://www.w3.org/2000/svg', 'g');
      grid.setAttribute('stroke', '#dde3ea');
      grid.setAttribute('stroke-width', '1');
      for (let i = 0; i <= 4; i++) {{
        const y = 40 + i * 78;
        const line = document.createElementNS('http://www.w3.org/2000/svg', 'line');
        line.setAttribute('x1', '64'); line.setAttribute('x2', '946');
        line.setAttribute('y1', y); line.setAttribute('y2', y);
        grid.appendChild(line);
      }}
      svg.appendChild(grid);
      const axis = document.createElementNS('http://www.w3.org/2000/svg', 'path');
      axis.setAttribute('d', 'M64 40 V352 H946');
      axis.setAttribute('fill', 'none');
      axis.setAttribute('stroke', '#9aa7b4');
      svg.appendChild(axis);
      const text = document.createElementNS('http://www.w3.org/2000/svg', 'g');
      text.setAttribute('font-size', '18');
      text.setAttribute('fill', '#627082');
      [[70, 384, labels[0]], [470, 384, labels[Math.floor(labels.length / 2)]], [850, 384, labels[labels.length - 1]]].forEach(([x,y,t]) => {{
        const node = document.createElementNS('http://www.w3.org/2000/svg', 'text');
        node.setAttribute('x', x); node.setAttribute('y', y); node.textContent = t || '';
        text.appendChild(node);
      }});
      const left = document.createElementNS('http://www.w3.org/2000/svg', 'text');
      left.setAttribute('x', 64); left.setAttribute('y', 25); left.textContent = leftTitle;
      text.appendChild(left);
      const right = document.createElementNS('http://www.w3.org/2000/svg', 'text');
      right.setAttribute('x', 760); right.setAttribute('y', 25); right.textContent = rightTitle;
      text.appendChild(right);
      svg.appendChild(text);
    }}
    function drawDualLine(containerId, rows, xLabel, leftKey, rightKey, leftLabel, rightLabel, rightFormatter = val) {{
      const svg = makeSvg(containerId);
      const n = rows.length;
      drawAxes(svg, rows.map(r => r[xLabel]), leftLabel, rightLabel);
      const [rankMin, rankMax] = metricRange(rows.map(r => r[leftKey]));
      const [rightMin, rightMax] = metricRange(rows.map(r => r[rightKey]));
      const x = i => 64 + (n <= 1 ? 0 : i * (882 / (n - 1)));
      const yRank = v => 40 + ((v - rankMin) / (rankMax - rankMin)) * 312;
      const yRight = v => 352 - ((v - rightMin) / (rightMax - rightMin)) * 312;
      const rankPts = rows.map((r, i) => ({{ x: x(i), y: yRank(r[leftKey]), row: r }}));
      const rightPts = rows.map((r, i) => ({{ x: x(i), y: yRight(r[rightKey]), row: r }}));
      [[rankPts, '#1464d2', 3], [rightPts, '#d65f32', 3]].forEach(([pts, color, width]) => {{
        const path = document.createElementNS('http://www.w3.org/2000/svg', 'path');
        path.setAttribute('d', linePath(pts));
        path.setAttribute('fill', 'none');
        path.setAttribute('stroke', color);
        path.setAttribute('stroke-width', width);
        svg.appendChild(path);
      }});
      rankPts.forEach((p, i) => {{
        const c = document.createElementNS('http://www.w3.org/2000/svg', 'circle');
        c.setAttribute('cx', p.x); c.setAttribute('cy', p.y); c.setAttribute('r', '5');
        c.setAttribute('fill', '#1464d2');
        c.addEventListener('mousemove', e => showTip(`${{rows[i][xLabel]}}<br>搜索排名: ${{fmt.format(rows[i][leftKey])}}<br>${{rightLabel}}: ${{rightFormatter(rows[i][rightKey])}}`, e));
        c.addEventListener('mouseleave', hideTip);
        svg.appendChild(c);
      }});
    }}
    function drawBars(containerId, rows, labelKey, valueKey, color = '#16845b', formatter = val) {{
      const svg = makeSvg(containerId);
      const max = Math.max(1, ...rows.map(r => r[valueKey] || 0));
      const barH = Math.min(34, 300 / Math.max(1, rows.length));
      rows.forEach((r, i) => {{
        const y = 42 + i * (barH + 8);
        const w = ((r[valueKey] || 0) / max) * 690;
        const label = document.createElementNS('http://www.w3.org/2000/svg', 'text');
        label.setAttribute('x', '28'); label.setAttribute('y', y + barH * .7);
        label.setAttribute('font-size', '16'); label.setAttribute('fill', '#17212b');
        label.textContent = r[labelKey];
        svg.appendChild(label);
        const rect = document.createElementNS('http://www.w3.org/2000/svg', 'rect');
        rect.setAttribute('x', '220'); rect.setAttribute('y', y);
        rect.setAttribute('width', Math.max(1, w)); rect.setAttribute('height', barH);
        rect.setAttribute('rx', '4'); rect.setAttribute('fill', r.isAbaTop ? '#d65f32' : color);
        rect.addEventListener('mousemove', e => showTip(`${{r[labelKey]}}<br>${{formatter(r[valueKey])}}${{r.isAbaTop ? '<br>ABA 前三 ASIN' : ''}}`, e));
        rect.addEventListener('mouseleave', hideTip);
        svg.appendChild(rect);
        const value = document.createElementNS('http://www.w3.org/2000/svg', 'text');
        value.setAttribute('x', 228 + w); value.setAttribute('y', y + barH * .7);
        value.setAttribute('font-size', '16'); value.setAttribute('fill', '#627082');
        value.textContent = formatter(r[valueKey]);
        svg.appendChild(value);
      }});
    }}

    function renderKpis() {{
      const latestTrend = DATA.trend[DATA.trend.length - 1];
      const firstTrend = DATA.trend[0];
      const latestMonth = [...new Set(DATA.monthly.map(r => r.month))].sort().pop();
      const latestMonthRows = DATA.monthly.filter(r => r.month === latestMonth);
      const latestSales = latestMonthRows.reduce((s, r) => s + r.sales, 0);
      const topProduct = DATA.products.slice().sort((a,b) => b.latestMonthSales - a.latestMonthSales)[0];
      const rankChange = firstTrend && latestTrend ? firstTrend.rank - latestTrend.rank : null;
      document.getElementById('meta').innerHTML =
        `来源：${{DATA.abaFile}}；Sales 文件 ${{DATA.products.length}} 个；关键词：<b>${{DATA.keyword}}</b>；数据区间：${{DATA.dateRange}}`;
      const kpis = [
        ['最新 ABA 排名', latestTrend ? fmt.format(latestTrend.rank) : '—', latestTrend ? latestTrend.end : ''],
        ['排名变化', rankChange == null ? '—' : (rankChange >= 0 ? '+' : '') + fmt.format(rankChange), '相对最早一周，正数代表排名改善'],
        ['最新月总销量', fmt.format(Math.round(latestSales)), latestMonth || ''],
        ['最高月销量 ASIN', topProduct ? topProduct.asin : '—', topProduct ? `${{fmt.format(Math.round(topProduct.latestMonthSales))}} 件` : ''],
        ['排名/销量相关', DATA.stats.rankSalesCorr == null ? '—' : DATA.stats.rankSalesCorr, '周排名数值 vs 周销量；负数代表排名越好销量越高'],
      ];
      document.getElementById('kpis').innerHTML = kpis.map(k => `<div class="kpi"><div class="label">${{k[0]}}</div><div class="value">${{k[1]}}</div><div class="sub">${{k[2]}}</div></div>`).join('');
    }}
    function renderControls() {{
      const months = [...new Set(DATA.monthly.map(r => r.month))].sort().reverse();
      document.getElementById('rankMonth').innerHTML = months.map(m => `<option value="${{m}}">${{m}}</option>`).join('');
      const productOpts = DATA.products.slice().sort((a,b) => b.latestMonthSales - a.latestMonthSales)
        .map(p => `<option value="${{p.asin}}">${{p.asin}} · ${{p.file.replace('.xlsx','')}}</option>`).join('');
      document.getElementById('asinSelect').innerHTML = productOpts;
    }}
    function renderCharts() {{
      drawDualLine('weeklyChart', DATA.weekly.filter(r => r.totalSales > 0), 'end', 'rank', 'totalSales', '搜索排名', '周销量');
      const topRows = DATA.topAsins.slice(0, 12).map(r => ({{...r, label: r.asin, value: r.top3Periods}}));
      drawBars('topAsinChart', topRows, 'label', 'value', '#1464d2', val);
      const metric = document.getElementById('monthlyMetric').value;
      drawDualLine('monthlyChart', DATA.monthlyComparison.filter(r => r.avgRank != null && r[metric] > 0), 'month', 'avgRank', metric, '月均搜索排名', metric === 'sales' ? '月销量' : '月销售额', metric === 'sales' ? val : money);
      renderAsinBar();
      renderAsinWeekly();
    }}
    function renderAsinBar() {{
      const month = document.getElementById('rankMonth').value;
      const topSet = new Set(DATA.topAsins.filter(t => t.latestRank != null || t.top3Periods > 0).map(t => t.asin));
      const rows = DATA.monthly.filter(r => r.month === month).sort((a,b) => b.sales - a.sales).slice(0, 14)
        .map(r => ({{ asin: r.asin, sales: r.sales, isAbaTop: topSet.has(r.asin) }}));
      drawBars('asinBarChart', rows, 'asin', 'sales', '#16845b', val);
    }}
    function renderAsinWeekly() {{
      const asin = document.getElementById('asinSelect').value;
      const rows = DATA.weekly.map(r => ({{ end: r.end, rank: r.rank, sales: r.asinSales[asin] || 0 }})).filter(r => r.sales > 0);
      drawDualLine('asinWeeklyChart', rows, 'end', 'rank', 'sales', '搜索排名', `${{asin}} 周销量`);
    }}
    function renderTable() {{
      const q = (document.getElementById('productFilter').value || '').trim().toLowerCase();
      const topMap = new Map(DATA.topAsins.map(t => [t.asin, t]));
      const rows = DATA.products
        .filter(p => !q || p.asin.toLowerCase().includes(q) || p.file.toLowerCase().includes(q))
        .sort((a,b) => b.latestMonthSales - a.latestMonthSales);
      document.getElementById('productTable').innerHTML = `
        <thead><tr><th>ASIN</th><th>文件</th><th>最新月</th><th>月销量</th><th>月销售额</th><th>ABA 占位</th><th>标题</th></tr></thead>
        <tbody>${{rows.map(p => {{
          const t = topMap.get(p.asin);
          const aba = t ? `${{t.latestRank ? '当前第 ' + t.latestRank : '未进前三'}} / 前三 ${{t.top3Periods}} 期` : '—';
          return `<tr><td><b>${{p.asin}}</b></td><td>${{p.file}}</td><td>${{p.latestMonth || '—'}}</td><td>${{val(p.latestMonthSales)}}</td><td>${{money.format(p.latestMonthRevenue || 0)}}</td><td>${{t ? '<span class="tag">' + aba + '</span>' : aba}}</td><td>${{p.title || ''}}</td></tr>`;
        }}).join('')}}</tbody>`;
    }}
    renderKpis();
    renderControls();
    renderCharts();
    renderTable();
    document.getElementById('monthlyMetric').addEventListener('change', renderCharts);
    document.getElementById('rankMonth').addEventListener('change', renderAsinBar);
    document.getElementById('asinSelect').addEventListener('change', renderAsinWeekly);
    document.getElementById('productFilter').addEventListener('input', renderTable);
    window.addEventListener('resize', renderCharts);
  </script>
</body>
</html>
"""


def main():
    aba_file, trend_rows, top_asins = load_aba()
    products, monthly_rows, daily_rows = load_sales()
    weekly = aggregate_weekly_sales(trend_rows, daily_rows)
    monthly_comparison = aggregate_monthly(monthly_rows, trend_rows)
    keyword = trend_rows[0]["keyword"] if trend_rows else ""
    date_range = f"{trend_rows[0]['start']} 至 {trend_rows[-1]['end']}" if trend_rows else ""
    stats = {"rankSalesCorr": correlation(weekly, "rank", "totalSales")}
    data = {
        "abaFile": aba_file,
        "keyword": keyword,
        "dateRange": date_range,
        "trend": trend_rows,
        "topAsins": top_asins,
        "products": products,
        "monthly": sorted(monthly_rows, key=lambda x: (x["month"], x["asin"])),
        "weekly": weekly,
        "monthlyComparison": monthly_comparison,
        "stats": stats,
    }
    OUTPUT_DIR.mkdir(exist_ok=True)
    OUTPUT_FILE.write_text(html_template(data), encoding="utf-8")
    print(OUTPUT_FILE)


if __name__ == "__main__":
    main()
